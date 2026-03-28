"""
RCSI News Digest — River Capital
Discovers investment-relevant news, summarises paywalled articles via Claude API,
sends an HTML email digest, and appends to a persistent Excel news log.
"""

import os
import re
import json
import smtplib
import hashlib
import base64
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import feedparser
import requests
from bs4 import BeautifulSoup
from anthropic import Anthropic
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

# ── Configuration ────────────────────────────────────────────────────────────

# Sources where ALL articles are auto-included (no keyword filtering)
# These publications are so consistently relevant that filtering drops good articles
AUTO_INCLUDE_SOURCES = ["QSR Media", "Franchise Business", "Inside Franchise Business"]
AUTO_INCLUDE_RSS = [
    "https://www.afr.com/rss/street-talk",  # AFR Street Talk — deal flow, PE, M&A
]

TOPICS = {
    "QSR & Franchising": [
        "QSR", "quick service restaurant", "fast food", "franchise",
        "franchisee", "franchising", "drive-thru", "food chain",
        "restaurant chain", "fast casual", "takeaway", "food court",
        # Australian QSR brands
        "Grill'd", "Betty's Burgers", "GYG", "Guzman y Gomez",
        "Boost Juice", "Schnitz", "Nando's", "Zambrero",
        "Oporto", "Red Rooster", "Chicken Treat", "Craveable",
        "Domino's", "Hungry Jack", "KFC", "McDonald's",
        "Subway", "Pizza Hut", "Pizza Capers",
        "Retail Food Group", "RFG", "Gloria Jean", "Brumby's",
        "Michel's Patisserie", "Donut King", "Crust Pizza",
        "Collins Foods", "Sushi Sushi", "Genki",
        "YOMG", "Augustus Gelatery", "Cheesecake Shop",
        "Chatime", "Gong Cha", "Starbucks Australia",
        "Noodle Box", "Roll'd", "San Churro",
        "Rashays", "Ribs & Burgers", "Zeus Street Greek",
        "Motto Motto", "Fishbowl", "Gami Chicken",
        "Taco Bell Australia", "Wendy's Australia", "Carl's Jr",
        "Muffin Break", "Jamaica Blue", "Foodco",
        "Oliver's Real Food", "Sumo Salad", "Soul Origin",
        "Mad Mex", "Fonda", "Hog's Breath",
    ],
    "Food & Beverage": [
        "brewery", "brewer", "beer", "craft beer", "beverage",
        "winery", "wine", "spirits", "distillery", "drinks",
        "hospitality", "pub", "hotel group", "bar chain",
        "cafe chain", "coffee chain", "bakery chain",
        "food manufacturer", "food processing", "FMCG food",
        "catering", "meal kit", "food delivery",
        # Australian F&B brands & companies
        "Lion", "Carlton & United", "CUB", "Coopers",
        "Treasury Wine", "Penfolds", "Accolade Wines",
        "Better Beer", "Inspired Unemployed",
        "Patties Foods", "Bega Cheese", "a2 Milk",
        "Goodman Fielder", "Inghams", "Tassal",
        "T2 Tea", "Aeropress", "Vittoria Coffee",
    ],
    "Private Equity & M&A": [
        "private equity", "PE deal", "leveraged buyout", "LBO",
        "acquisition", "divestiture", "portfolio company", "bolt-on",
        "add-on acquisition", "management buyout", "MBO",
        "sponsor-backed", "take-private", "exit multiple",
        "dry powder", "fund raise", "capital raise", "capital raising",
        "hires CEO", "new CEO", "appoints CEO", "names CEO",
        "management change", "board appointment", "activist investor",
        "shareholder", "stake", "strategic review", "recapitalisation",
        "IPO", "listing", "float",
        # Deal language (headlines)
        "circles", "eyes", "targets", "swoops", "bids for",
        "in talks", "weighs sale", "explores sale", "mandate",
        "due diligence", "binding offer", "indicative offer",
        "scheme of arrangement", "merger", "demerger",
        "receivership", "administration", "collapsed",
        "Luminis Partners", "Kroll", "Greenhill",
        # PE firms — Australian
        "KKR", "Bain Capital", "Pacific Equity Partners",
        "BGH Capital", "Quadrant", "Advent Partners",
        "Allegro Funds", "Next Capital", "Adamantem",
        "Archer Capital", "Crescent Capital", "Navis Capital",
        "Anchorage Capital", "Five V Capital",
        "Tattarang", "Andrew Forrest", "CHAMP",
        "Wolseley Private Equity", "Mercury Capital",
        "Pemba Capital", "Ironbridge",
        # PE firms — Global (active in AU)
        "Partners Group", "Blackstone", "Carlyle",
        "Apollo", "Warburg Pincus", "TPG Capital",
        "CVC Capital", "EQT Partners", "Permira",
        "Advent International", "Cinven", "PAG",
        # Columns & sections
        "Street Talk", "DataRoom", "Rear Window",
    ],
    "Retail & Consumer": [
        "retail", "consumer spending", "discretionary",
        "same-store sales", "SSS", "like-for-like",
        "foot traffic", "consumer sentiment", "FMCG",
        "shopping centre", "retail sales", "consumer confidence",
        "pharmacy", "chemist", "Priceline", "Chemist Warehouse",
        "Woolworths", "Coles", "Wesfarmers", "JB Hi-Fi",
        "Kmart", "Target Australia", "Bunnings",
        "Premier Investments", "Lovisa", "Cotton On",
        "Country Road", "David Jones", "Myer",
        "Harvey Norman", "Super Retail", "Rebel Sport",
        "Accent Group", "Adairs", "Baby Bunting",
        "Endeavour Group", "BWS", "Dan Murphy",
        "Sigma Healthcare", "API", "Infinity",
    ],
    "Australian Economy & Markets": [
        "RBA", "interest rate", "Reserve Bank", "inflation Australia",
        "GDP Australia", "unemployment Australia", "ASX",
        "Australian dollar", "AUD", "fiscal policy Australia",
        "federal budget", "cost of living",
    ],
}

REPUTABLE_SOURCES = {
    "afr.com": "AFR",
    "theaustralian.com.au": "The Australian",
    "reuters.com": "Reuters",
    "bloomberg.com": "Bloomberg",
    "ft.com": "Financial Times",
    "wsj.com": "Wall Street Journal",
    "qsrmedia.com.au": "QSR Media",
    "franchisebusiness.com.au": "Franchise Business",
    "insidefranchisebusiness.com.au": "Inside Franchise Business",
    "smartcompany.com.au": "SmartCompany",
    "businessnewsaustralia.com": "Business News Australia",
    "abc.net.au": "ABC News",
    "smh.com.au": "SMH",
    "theage.com.au": "The Age",
    "news.com.au": "News.com.au",
    "9news.com.au": "Nine News",
    "theguardian.com": "The Guardian",
}

PAYWALL_DOMAINS = ["afr.com", "theaustralian.com.au"]

RSS_FEEDS = [
    "https://www.afr.com/rss/companies",
    "https://www.afr.com/rss/street-talk",
    "https://www.afr.com/rss/markets",
    "https://www.afr.com/rss/policy",
    "https://www.theaustralian.com.au/feed",
    "https://www.qsrmedia.com.au/feed",
]

GOOGLE_NEWS_QUERIES = [
    # ── AFR Street Talk (broad — catches all deal activity) ──
    "AFR Street Talk",
    "site:afr.com street-talk",
    # ── QSR brand-specific ──
    "Craveable Brands OR Oporto OR Red Rooster",
    "Retail Food Group OR Gloria Jean's OR Donut King",
    "Grill'd OR Betty's Burgers OR Zambrero",
    "Guzman y Gomez OR GYG Australia",
    "Collins Foods OR Domino's Australia",
    "Sushi Sushi OR Genki Global OR Roll'd",
    "Boost Juice OR Chatime OR Gong Cha",
    "YOMG OR Augustus Gelatery OR Cheesecake Shop",
    # ── Food & Beverage ──
    "Australian brewery OR craft beer capital raising",
    "Australia food beverage acquisition OR investment",
    "Better Beer OR Inspired Unemployed",
    # ── PE & M&A (broad) ──
    "private equity Australia acquisition",
    "private equity Australia food OR restaurant OR franchise OR retail",
    "Partners Group OR BGH Capital OR Quadrant Australia",
    "Allegro Funds OR Adamantem OR Anchorage Capital Australia",
    # ── Theme-based (broad) ──
    "QSR franchise Australia",
    "restaurant chain Australia CEO OR acquisition OR expansion",
    "franchise Australia sale OR acquisition OR investor",
    "Australian fast food chain",
    "Australia retail pharmacy OR Priceline OR Chemist Warehouse deal",
]

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT") or "587")
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASS = os.getenv("SMTP_PASS")
EMAIL_TO = os.getenv("EMAIL_TO")
EMAIL_FROM = os.getenv("EMAIL_FROM", SMTP_USER)

AFR_EMAIL = os.getenv("AFR_EMAIL")
AFR_PASSWORD = os.getenv("AFR_PASSWORD")
AUSTRALIAN_EMAIL = os.getenv("AUSTRALIAN_EMAIL")
AUSTRALIAN_PASSWORD = os.getenv("AUSTRALIAN_PASSWORD")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")

EXCEL_PATH = Path(os.getenv("EXCEL_PATH", "news_log.xlsx"))

# ── Helpers ──────────────────────────────────────────────────────────────────

def get_lookback_hours():
    """72 hours on Monday, 24 hours otherwise."""
    now = datetime.now(timezone.utc)
    return 72 if now.weekday() == 0 else 24


def is_reputable(url):
    for domain in REPUTABLE_SOURCES:
        if domain in url:
            return True, REPUTABLE_SOURCES[domain]
    return False, None


def classify_article(title, snippet=""):
    """Return the first matching topic, or None."""
    text = f"{title} {snippet}".lower()
    for topic, keywords in TOPICS.items():
        for kw in keywords:
            if kw.lower() in text:
                return topic
    return None


def article_id(url):
    return hashlib.md5(url.encode()).hexdigest()


def decode_google_news_url(google_url):
    """Decode the actual article URL from a Google News encoded URL."""
    match = re.search(r'/articles/(.+?)(?:\?|$)', google_url)
    if not match:
        return None
    article_id_str = match.group(1)
    # Pad base64 if needed
    padding = 4 - len(article_id_str) % 4
    if padding != 4:
        article_id_str += '=' * padding
    try:
        decoded = base64.urlsafe_b64decode(article_id_str)
        urls = re.findall(rb'https?://[^\s\x00-\x1f"\'<>]+', decoded)
        for url_bytes in urls:
            url_str = url_bytes.decode('utf-8', errors='ignore')
            return url_str
    except Exception:
        pass
    return None


def is_auto_include_rss(feed_url):
    """Check if this RSS feed should auto-include all articles."""
    return any(auto_url in feed_url for auto_url in AUTO_INCLUDE_RSS)


# ── Google News Source Matching ──────────────────────────────────────────────

SOURCE_NAME_MAP = {
    "australian financial review": "AFR",
    "afr": "AFR",
    "the australian": "The Australian",
    "theaustralian.com.au": "The Australian",
    "reuters": "Reuters",
    "bloomberg": "Bloomberg",
    "financial times": "Financial Times",
    "wall street journal": "Wall Street Journal",
    "wsj": "Wall Street Journal",
    "qsr media": "QSR Media",
    "qsrmedia": "QSR Media",
    "qsrmedia.com.au": "QSR Media",
    "franchise business": "Franchise Business",
    "inside franchise business": "Inside Franchise Business",
    "smartcompany": "SmartCompany",
    "business news australia": "Business News Australia",
    "abc news": "ABC News",
    "abc": "ABC News",
    "sydney morning herald": "SMH",
    "smh": "SMH",
    "the age": "The Age",
    "news.com.au": "News.com.au",
    "nine news": "Nine News",
    "9news": "Nine News",
    "9news.com.au": "Nine News",
    "the guardian": "The Guardian",
    "guardian australia": "The Guardian",
}


# ── Article Discovery ────────────────────────────────────────────────────────

def fetch_rss_articles(cutoff):
    articles = {}
    for feed_url in RSS_FEEDS:
        try:
            auto_include = is_auto_include_rss(feed_url)
            feed = feedparser.parse(feed_url)
            entry_count = len(feed.entries) if feed.entries else 0
            print(f"[RSS] {feed_url} — {entry_count} entries{' (auto-include)' if auto_include else ''}")
            for entry in feed.entries:
                url = entry.get("link", "")
                if not url:
                    continue
                reputable, source_name = is_reputable(url)
                if not reputable:
                    continue
                published = None
                if hasattr(entry, "published_parsed") and entry.published_parsed:
                    from calendar import timegm
                    published = datetime.fromtimestamp(timegm(entry.published_parsed), tz=timezone.utc)
                if published and published < cutoff:
                    continue
                title = entry.get("title", "").strip()
                snippet = entry.get("summary", "").strip()

                # Auto-include sources skip keyword filtering
                if auto_include:
                    topic = classify_article(title, snippet) or "Private Equity & M&A"
                else:
                    topic = classify_article(title, snippet)
                    if not topic:
                        continue

                aid = article_id(url)
                if aid not in articles:
                    articles[aid] = {
                        "title": title,
                        "url": url,
                        "source": source_name,
                        "topic": topic,
                        "date": published or datetime.now(timezone.utc),
                        "snippet": snippet,
                    }
        except Exception as e:
            print(f"[RSS] Error fetching {feed_url}: {e}")
    print(f"[RSS] Total: {len(articles)} articles from RSS feeds")
    return articles


def fetch_google_news(cutoff):
    articles = {}
    skipped_sources = {}

    for query in GOOGLE_NEWS_QUERIES:
        try:
            url = f"https://news.google.com/rss/search?q={requests.utils.quote(query)}+when:7d&hl=en-AU&gl=AU&ceid=AU:en"
            feed = feedparser.parse(url)
            entry_count = len(feed.entries) if feed.entries else 0
            print(f"[Google News] '{query}' — {entry_count} entries")
            for entry in feed.entries:
                link = entry.get("link", "")
                if not link:
                    continue

                # Use the source tag from the feed (no URL resolution needed)
                source_tag = entry.get("source", {})
                if isinstance(source_tag, dict):
                    source_text = source_tag.get("title", "")
                elif hasattr(source_tag, "title"):
                    source_text = source_tag.title
                else:
                    source_text = str(source_tag)

                source_name = SOURCE_NAME_MAP.get(source_text.lower().strip())
                if not source_name:
                    for key, name in SOURCE_NAME_MAP.items():
                        if key in source_text.lower():
                            source_name = name
                            break
                if not source_name:
                    skipped_sources[source_text] = skipped_sources.get(source_text, 0) + 1
                    continue

                published = None
                if hasattr(entry, "published_parsed") and entry.published_parsed:
                    from calendar import timegm
                    published = datetime.fromtimestamp(timegm(entry.published_parsed), tz=timezone.utc)
                if published and published < cutoff:
                    continue

                title = entry.get("title", "").strip()
                title = re.sub(r"\s*-\s*[^-]+$", "", title)

                # Auto-include for QSR Media, Franchise Business etc.
                is_auto = source_name in AUTO_INCLUDE_SOURCES
                if is_auto:
                    topic = classify_article(title) or "QSR & Franchising"
                else:
                    topic = classify_article(title)
                    if not topic:
                        topic = classify_article(title, entry.get("summary", ""))
                    if not topic:
                        continue

                aid = article_id(link)
                if aid not in articles:
                    articles[aid] = {
                        "title": title,
                        "url": link,
                        "source": source_name,
                        "topic": topic,
                        "date": published or datetime.now(timezone.utc),
                        "snippet": "",
                    }
        except Exception as e:
            print(f"[Google News] Error with query '{query}': {e}")

    if skipped_sources:
        top_skipped = sorted(skipped_sources.items(), key=lambda x: -x[1])[:10]
        print(f"[Google News] Top skipped sources: {top_skipped}")
    print(f"[Google News] Total: {len(articles)} articles from Google News")
    return articles


# ── Paywall Authentication & Scraping ────────────────────────────────────────

def login_afr(session):
    try:
        resp = session.post(
            "https://login.nine.com.au/api/login",
            json={"email": AFR_EMAIL, "password": AFR_PASSWORD, "client_id": "afr"},
            headers={
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Origin": "https://www.afr.com",
                "Referer": "https://www.afr.com/",
            },
            timeout=30,
        )
        if resp.status_code == 200:
            print("[AFR] Login successful")
            return True
        print(f"[AFR] Login failed: {resp.status_code} — {resp.text[:200]}")
        return False
    except Exception as e:
        print(f"[AFR] Login error: {e}")
        return False


def login_australian(session):
    # Try multiple known News Corp auth endpoints
    endpoints = [
        "https://api.newscorpaustralia.com/v4/login",
        "https://auth-api.news.com.au/v4/login",
        "https://component-api.news.com.au/v4/login",
    ]
    for login_url in endpoints:
        try:
            resp = session.post(
                login_url,
                json={"email": AUSTRALIAN_EMAIL, "password": AUSTRALIAN_PASSWORD},
                headers={
                    "Content-Type": "application/json",
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                    "Origin": "https://www.theaustralian.com.au",
                    "Referer": "https://www.theaustralian.com.au/",
                },
                timeout=15,
            )
            if resp.status_code == 200:
                print(f"[The Australian] Login successful via {login_url}")
                return True
            print(f"[The Australian] {login_url} — {resp.status_code}")
        except Exception as e:
            print(f"[The Australian] {login_url} — {e}")
    print("[The Australian] All login endpoints failed")
    return False


def scrape_afr_sections(session, cutoff):
    """Scrape article links directly from AFR section pages using authenticated session."""
    AFR_SECTION_URLS = [
        "https://www.afr.com/street-talk",
        "https://www.afr.com/companies",
        "https://www.afr.com/markets",
    ]
    articles = {}
    for section_url in AFR_SECTION_URLS:
        try:
            resp = session.get(section_url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            }, timeout=30)
            if resp.status_code != 200:
                print(f"[AFR Scrape] Non-200 for {section_url}: {resp.status_code}")
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
            section_articles = {}
            for a_tag in soup.find_all("a", href=True):
                href = a_tag["href"]
                if href.startswith("/"):
                    href = f"https://www.afr.com{href}"
                if "afr.com" not in href:
                    continue
                # Skip non-article links
                skip = ["/topic/", "/by/", "/author/", "/video/", "/podcast/",
                        "/newsletters", "/subscribe", "/login", "/search",
                        "javascript:", "/rss", "/accessibility"]
                if any(p in href.lower() for p in skip):
                    continue
                # Articles must have a slug with an ID pattern (e.g. -p5pt7k)
                if not re.search(r'-p[a-z0-9]{4,}$', href.rstrip('/')):
                    # Or a date-like pattern
                    if not re.search(r'/\d{8}-', href):
                        continue
                title = a_tag.get_text(strip=True)
                if not title or len(title) < 15:
                    continue
                # Dedup within section
                if href in section_articles:
                    continue
                section_articles[href] = title

            for href, title in section_articles.items():
                topic = classify_article(title)
                if not topic and "street-talk" in section_url:
                    topic = "Private Equity & M&A"
                if not topic:
                    continue
                aid = article_id(href)
                if aid not in articles:
                    articles[aid] = {
                        "title": title,
                        "url": href,
                        "source": "AFR",
                        "topic": topic,
                        "date": datetime.now(timezone.utc),
                        "snippet": "",
                    }
            print(f"[AFR Scrape] {section_url} — {len(section_articles)} article links found")
        except Exception as e:
            print(f"[AFR Scrape] Error scraping {section_url}: {e}")
    print(f"[AFR Scrape] Total: {len(articles)} relevant articles from AFR sections")
    return articles


def fetch_full_article(url, session):
    try:
        resp = session.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }, timeout=30)
        if resp.status_code != 200:
            print(f"[Scrape] Non-200 for {url}: {resp.status_code}")
            return None
        soup = BeautifulSoup(resp.text, "html.parser")

        # AFR uses hashed CSS module classes like "b661c1fd082d4a03-articleBody"
        body = soup.find("div", class_=re.compile(r'articleBody', re.I))
        if not body:
            body = soup.find("div", {"id": "article-body"})
        if not body:
            body = soup.find("div", class_=re.compile(r'article[_-]?body|story[_-]?body', re.I))
        if not body:
            body = soup.find("div", {"id": "story"})
        if not body:
            body = soup.find("article")

        if body:
            text = "\n".join(p.get_text(strip=True) for p in body.find_all("p") if p.get_text(strip=True))
            if len(text) > 200:
                return text

        # Fallback: grab all paragraphs from the page if no container found
        all_paras = soup.find_all("p")
        text = "\n".join(p.get_text(strip=True) for p in all_paras if len(p.get_text(strip=True)) > 40)
        if len(text) > 200:
            print(f"[Scrape] Used fallback paragraph extraction for {url[:60]}")
            return text

        print(f"[Scrape] Could not extract body from {url[:60]}")
        return None
    except Exception as e:
        print(f"[Scrape] Error fetching {url[:60]}: {e}")
        return None


# ── Summarisation ────────────────────────────────────────────────────────────

def summarise_article(title, full_text):
    if not ANTHROPIC_API_KEY or not full_text:
        return ""
    try:
        client = Anthropic(api_key=ANTHROPIC_API_KEY)
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=300,
            messages=[{"role": "user", "content": (
                "You are an investment analyst at an Australian mid-market private equity firm "
                "focused on QSR and franchise investments.\n\n"
                "Summarise the following article in 2-3 concise sentences. Focus on what is "
                "most relevant for PE deal origination, valuation, and portfolio monitoring.\n\n"
                f"Article title: {title}\n\nArticle text:\n{full_text[:12000]}"
            )}],
        )
        return response.content[0].text.strip()
    except Exception as e:
        print(f"[Summary] Error summarising '{title[:50]}': {e}")
        return ""


# ── Excel News Log ───────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
HEADERS = ["Date", "Headline", "Source", "URL", "Topic", "Summary"]
COL_WIDTHS = [14, 55, 22, 50, 28, 65]


def init_excel(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "News Log"
    ws.freeze_panes = "A2"
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.auto_filter.ref = "A1:F1"
    wb.save(path)
    return wb


def append_to_excel(articles_list):
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb["News Log"]
    else:
        wb = init_excel(EXCEL_PATH)
        ws = wb["News Log"]

    existing_urls = set()
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        if row[0]:
            existing_urls.add(row[0].strip())

    new_count = 0
    for art in articles_list:
        if art["url"].strip() in existing_urls:
            continue
        next_row = ws.max_row + 1
        date_str = art["date"].strftime("%Y-%m-%d") if isinstance(art["date"], datetime) else str(art["date"])
        values = [date_str, art["title"], art["source"], art["url"], art["topic"], art.get("summary", "")]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in [2, 6]))
        new_count += 1

    wb.save(EXCEL_PATH)
    print(f"[Excel] Appended {new_count} new articles to {EXCEL_PATH}")
    return new_count


# ── Email Digest ─────────────────────────────────────────────────────────────

def build_email_html(articles_by_topic, run_date):
    lookback = get_lookback_hours()
    html = f"""
    <html><body style="font-family: Arial, sans-serif; color: #1a1a1a; max-width: 700px; margin: 0 auto;">
    <div style="background: #1B2A4A; color: white; padding: 20px 24px; border-radius: 6px 6px 0 0;">
        <h1 style="margin:0; font-size:20px;">RCSI News Digest</h1>
        <p style="margin:6px 0 0; font-size:13px; color:#a0b4d0;">
            {run_date.strftime('%A, %d %B %Y')} &nbsp;|&nbsp; Last {lookback} hours
        </p>
    </div>
    <div style="padding: 20px 24px; background: #f9f9fb; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 6px 6px;">
    """

    if not articles_by_topic:
        html += '<p style="color:#666;">No relevant articles found in this period.</p>'
    else:
        # Display topics in preferred order
        topic_order = [
            "QSR & Franchising", "Food & Beverage", "Private Equity & M&A",
            "Retail & Consumer", "Australian Economy & Markets",
        ]
        for topic in topic_order:
            arts = articles_by_topic.get(topic, [])
            if not arts:
                continue
            html += f'<h2 style="font-size:16px; color:#1B2A4A; border-bottom:2px solid #1B2A4A; padding-bottom:4px; margin-top:24px;">{topic} ({len(arts)})</h2>'
            for art in arts:
                summary_block = ""
                if art.get("summary"):
                    summary_block = f'<p style="margin:4px 0 0; font-size:13px; color:#444; line-height:1.5;">{art["summary"]}</p>'
                html += f"""
                <div style="margin-bottom:16px;">
                    <a href="{art['url']}" style="font-size:14px; color:#1B2A4A; font-weight:600; text-decoration:none;">
                        {art['title']}
                    </a>
                    <span style="font-size:12px; color:#888; margin-left:8px;">{art['source']}</span>
                    {summary_block}
                </div>
                """

    html += """
    <p style="font-size:11px; color:#999; margin-top:30px; border-top:1px solid #ddd; padding-top:10px;">
        STRICTLY PRIVATE &amp; CONFIDENTIAL — River Capital
    </p>
    </div></body></html>
    """
    return html


def send_email(html, run_date):
    if not all([SMTP_USER, SMTP_PASS, EMAIL_TO]):
        print("[Email] Missing SMTP credentials — skipping send")
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"RCSI News Digest — {run_date.strftime('%d %b %Y')}"
        msg["From"] = EMAIL_FROM
        msg["To"] = EMAIL_TO
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(EMAIL_FROM, [EMAIL_TO], msg.as_string())
        print("[Email] Digest sent successfully")
        return True
    except Exception as e:
        print(f"[Email] Send failed: {e}")
        return False


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    run_date = datetime.now(timezone.utc)
    lookback = get_lookback_hours()
    cutoff = run_date - timedelta(hours=lookback)
    print(f"[Run] {run_date.strftime('%Y-%m-%d %H:%M UTC')} — looking back {lookback}h to {cutoff.strftime('%Y-%m-%d %H:%M UTC')}")

    # 1. Authenticate first (needed for AFR section scraping)
    afr_session = requests.Session()
    aus_session = requests.Session()
    afr_logged_in = login_afr(afr_session) if AFR_EMAIL else False
    aus_logged_in = login_australian(aus_session) if AUSTRALIAN_EMAIL else False

    # 2. Discover articles from all sources
    articles = {}
    articles.update(fetch_rss_articles(cutoff))
    articles.update(fetch_google_news(cutoff))

    # 3. Scrape AFR section pages directly (gets real URLs, bypasses Google News encoding)
    if afr_logged_in:
        articles.update(scrape_afr_sections(afr_session, cutoff))

    print(f"[Discovery] {len(articles)} total articles found")

    if not articles:
        print("[Run] No articles — sending empty digest")
        html = build_email_html({}, run_date)
        send_email(html, run_date)
        return

    # 4. Fetch full text & summarise AFR articles (they have real URLs from scraping)
    for aid, art in articles.items():
        if art.get("source") != "AFR" or not afr_logged_in:
            continue
        # Skip Google News URLs — we can't resolve them
        if "news.google.com" in art["url"]:
            continue
        if "afr.com" not in art["url"]:
            continue
        full_text = fetch_full_article(art["url"], afr_session)
        if full_text:
            art["summary"] = summarise_article(art["title"], full_text)
            print(f"[Summary] Generated for: {art['title'][:60]}")
        else:
            art["summary"] = ""

    # 5. Group by topic
    articles_list = sorted(articles.values(), key=lambda a: a["date"], reverse=True)
    articles_by_topic = {}
    for art in articles_list:
        articles_by_topic.setdefault(art["topic"], []).append(art)

    # 6. Send email digest
    html = build_email_html(articles_by_topic, run_date)
    send_email(html, run_date)

    # 7. Append to Excel log
    append_to_excel(articles_list)

    print(f"[Run] Complete — {len(articles_list)} articles processed")


if __name__ == "__main__":
    main()
